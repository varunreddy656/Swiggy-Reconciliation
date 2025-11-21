import openpyxl
from pathlib import Path
import re
import shutil


# ----------------- Helper Functions -----------------

def extract_swiggy_start_day(filepath):
    """Extract numeric starting day from Summary!C12"""
    try:
        wb = openpyxl.load_workbook(filepath, data_only=True, read_only=True)
        sheet = wb["Summary"]
        text = str(sheet["C12"].value)
        wb.close()
    except:
        return None

    m = re.search(r"(\d+)\s*.*?[-to]+\s*(\d+)", text, re.IGNORECASE)
    if m:
        return int(m.group(1))
    return None


def detect_platform(fp):
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        sheets = wb.sheetnames
        wb.close()
    except:
        return None
    if "Other charges and deductions" in sheets:
        return "Swiggy"
    if "Addition Deductions Details" in sheets:
        return "Zomato"
    return None


def clear_all_D_sheets(wb):
    sheets_to_remove = [sh for sh in wb.sheetnames if sh.startswith("D1W") or sh.startswith("D2W")]
    for sh_name in sheets_to_remove:
        std = wb[sh_name]
        wb.remove(std)
    print(f"Cleared {len(sheets_to_remove)} old D1W/D2W sheets.")


def ensure_sheet(wb, name):
    if name in wb.sheetnames:
        return wb[name]
    else:
        return wb.create_sheet(name)


def copy_data(src, tgt, start_row):
    max_row, max_col = src.max_row, src.max_column
    tgt.delete_rows(1, tgt.max_row)
    for r in range(start_row, max_row + 1):
        for c in range(1, max_col + 1):
            tgt.cell(row=r - start_row + 1, column=c).value = src.cell(row=r, column=c).value


def extract_total_orders(fp):
    try:
        wb = openpyxl.load_workbook(fp, data_only=True, read_only=True)
        sheet = wb["Summary"]
        for row in range(1, sheet.max_row + 1):
            cell_value = sheet.cell(row=row, column=2).value
            if cell_value and "Total Orders" in str(cell_value) and (
                    "Delivered" in str(cell_value) or "Cancelled" in str(cell_value)):
                total_orders_value = sheet.cell(row=row, column=3).value
                wb.close()
                return total_orders_value
        wb.close()
        return None
    except Exception as e:
        print(f"Error extracting Total Orders from {fp.name}: {e}")
        return None


# ----------------- Calculations and Mapping -----------------
def map_values_to_cashflow(wb, data1_sheet, week):
    cashflow = wb["Cashflow"]
    week_col = 3 + (week - 1)

    data2_sheet_name = f"D2W{week}"
    data2_sheet = wb[data2_sheet_name] if data2_sheet_name in wb.sheetnames else None

    mapping = {
        "Item sales (Delivered orders)": (["Item Total"], 2, "single"),
        "Add:- Packing charges": (["Packaging Charges"], 2, "single"),
        "Add:- Compensation paid for cancelled orders": (
        ["Total Customer Paid", "Complaint & Cancellation Charges"], 1, "sub"),
        "Less:- Discount": (["Restaurant Discounts", "Swiggy One Exclusive Offer Discount"], 2, "sum"),
        "Add:- GST 5%": (["GST Collected"], 2, "single"),
        "Swiggy One Fees": (["Swiggy One Fees"], 3, "single"),
        "Call Center Service Fees": (["Call Center Charges"], 3, "single"),
        "PocketHero Fee": (["Pocket Hero Fees"], 3, "single"),
        "Platform Fee": (["Commission"], 3, "single"),
        "Long Distance Fee": (["Long Distance Charges"], 3, "single"),
        "Merchant Cancellation Charges": (["Restaurant Cancellation Charges"], 3, "single"),
        "Paid by Restaurant": (["Customer Complaints"], 4, "single"),
        "TDS deduction for aggrigators": (["TDS"], 4, "single"),
        "TCS": (["TCS"], 4, "single"),
        "GST collected and paid by swiggy": (["GST Deduction"], 4, "single"),
        "Collection Charges": (["Payment Collection Charges"], 3, "single")
    }

    partial_match_keywords = {
        "Total Customer Paid": "Total Customer Paid",
        "Complaint & Cancellation Charges": ["Complaint", "Cancellation"],
        "Restaurant Discounts": "Restaurant Discount",
        "Swiggy One Exclusive Offer Discount": "Swiggy One",
        "TCS": "TCS"
    }

    headers = {str(data1_sheet.cell(row=5, column=c).value).strip(): c
               for c in range(1, data1_sheet.max_column + 1)
               if data1_sheet.cell(row=5, column=c).value}

    def find_column(header_name):
        h_clean = header_name.strip()
        if h_clean in headers:
            return headers[h_clean]
        if h_clean in partial_match_keywords:
            keyword = partial_match_keywords[h_clean]
            if isinstance(keyword, list):
                for actual_header, col in headers.items():
                    if all(kw.lower() in actual_header.lower() for kw in keyword):
                        return col
            else:
                for actual_header, col in headers.items():
                    if keyword.lower() in actual_header.lower():
                        return col
        return None

    for row in range(1, cashflow.max_row + 1):
        label = cashflow.cell(row=row, column=2).value
        if not label:
            continue
        label = str(label).strip()
        if label not in mapping:
            continue

        data_headers, data_row, operation = mapping[label]
        data_cells = []

        for h in data_headers:
            col = find_column(h)
            if col:
                data_cells.append(data1_sheet.cell(row=data_row, column=col))
            else:
                print(f"Warning: Header '{h.strip()}' not found in Data1 sheet for '{label}'")

        if not data_cells:
            print(f"Skipping '{label}' because no Data1 headers found")
            continue

        formula = ""
        if operation == "single":
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}"
        elif operation == "sum":
            formula = "=" + "+".join([f"'{data1_sheet.title}'!{c.coordinate}" for c in data_cells])
        elif operation == "sub":
            if len(data_cells) != 2:
                print(f"Skipping subtraction for '{label}' because requires exactly 2 cells")
                continue
            formula = f"='{data1_sheet.title}'!{data_cells[0].coordinate}-'{data1_sheet.title}'!{data_cells[1].coordinate}"

        cashflow.cell(row=row, column=week_col).value = formula

    if data2_sheet:
        for row in range(1, cashflow.max_row + 1):
            label = cashflow.cell(row=row, column=2).value
            if not label:
                continue
            label = str(label).strip()
            if label == "High Priority":
                total_adj_row = None
                for r in range(1, data2_sheet.max_row + 1):
                    cell_value = data2_sheet.cell(row=r, column=1).value
                    if cell_value and "Total Adjustments" in str(cell_value):
                        total_adj_row = r
                        break
                if total_adj_row:
                    value_cell = data2_sheet.cell(row=total_adj_row, column=2)
                    formula = f"=-'{data2_sheet.title}'!{value_cell.coordinate}"
                    cashflow.cell(row=row, column=week_col).value = formula
                    print(f"High Priority mapped from {data2_sheet.title} row {total_adj_row}")
                else:
                    print(f"Warning: 'Total Adjustments' not found in {data2_sheet.title}")
                break
    else:
        print(f"Warning: Data2 sheet '{data2_sheet_name}' not found for week {week}")

    print(f"Cashflow mapped for week {week}")


def perform_calculations_on_data1(wb, data1_sheet, week, recon_path):
    data1_sheet.insert_rows(1, 4)

    item_total_col = None
    order_status_col = None
    for col_num, cell in enumerate(data1_sheet[5], 1):
        if cell.value == 'Item Total': item_total_col = col_num
        if cell.value == 'Order Status': order_status_col = col_num
    if not item_total_col or not order_status_col:
        print("Required columns missing")
        return

    delivered = [0] * (82 - item_total_col)
    cancelled = [0] * (82 - item_total_col)

    for row in range(6, data1_sheet.max_row + 1):
        status = str(data1_sheet.cell(row=row, column=order_status_col).value).strip().lower()
        if status not in ["delivered", "cancelled"]: continue
        target = delivered if status == "delivered" else cancelled
        for i, col in enumerate(range(item_total_col, 82)):
            val = data1_sheet.cell(row=row, column=col).value
            if isinstance(val, (int, float)): target[i] += val

    for i, col in enumerate(range(item_total_col, 82)):
        data1_sheet.cell(row=4, column=col).value = delivered[i] + cancelled[i]
        data1_sheet.cell(row=2, column=col).value = delivered[i]
        data1_sheet.cell(row=1, column=col).value = cancelled[i]

    for col in range(item_total_col, 82):
        val = data1_sheet.cell(row=4, column=col).value
        data1_sheet.cell(row=3, column=col).value = val * 1.18 if isinstance(val, (int, float)) else 0

    wb.save(recon_path)
    print("Row1/Row2/Row3/Row4 calculations done for", data1_sheet.title)

    map_values_to_cashflow(wb, data1_sheet, week)


# ----------------- Main Processing Function (Web Version) -----------------

def process_invoices_web(invoice_folder_path, template_recon_path, output_path, client_name=None):
    try:
        folder = Path(invoice_folder_path)
        shutil.copy2(template_recon_path, output_path)
        recon = openpyxl.load_workbook(output_path)

        clear_all_D_sheets(recon)

        invoice_files = list(folder.glob("*.xlsx"))
        invoices = []
        for fp in invoice_files:
            plat = detect_platform(fp)
            if plat == "Swiggy":
                d = extract_swiggy_start_day(fp)
                if d is not None:
                    invoices.append((d, fp, plat))
                else:
                    print(f"Skipping {fp.name}: Could not parse start day")

        if not invoices:
            return {
                'success': False,
                'message': 'No valid Swiggy invoices found. Please check your files.'
            }

        invoices.sort(key=lambda x: x[0])
        week_map = {}
        current_week = 1
        prev_day = None
        for day, fp, plat in invoices:
            if prev_day is None:
                week_map[fp] = current_week
            else:
                if day > prev_day:
                    current_week += 1
                week_map[fp] = current_week
            prev_day = day

        print("\nWeek map:")
        for fp, w in week_map.items():
            print(f"{fp.name} → Week {w}")

        summary_sheet = ensure_sheet(recon, "Summary")

        # New: write client name if given
        if client_name:
            summary_sheet.cell(row=1, column=2).value = client_name

        for d, fp, plat in invoices:
            week = week_map[fp]
            print(f"\nProcessing {fp.name} → Week {week}")
            wb_invoice = openpyxl.load_workbook(fp, data_only=True)
            ol = wb_invoice["Order Level"]
            add = wb_invoice["Other charges and deductions"]
            d1 = ensure_sheet(recon, f"D1W{week}")
            d2 = ensure_sheet(recon, f"D2W{week}")
            copy_data(ol, d1, 3)
            copy_data(add, d2, 4)
            total_orders = extract_total_orders(fp)
            if total_orders is not None:
                target_col = 2 + week
                summary_sheet.cell(row=6, column=target_col).value = total_orders
                print(f"Total Orders ({total_orders}) pasted in Summary sheet, Week {week}")
            else:
                print(f"Warning: Could not extract Total Orders from {fp.name}")
            wb_invoice.close()
            perform_calculations_on_data1(recon, d1, week, output_path)

        recon.save(output_path)
        print("\n✔ DONE — Data copied, calculations performed, and Cashflow mapped successfully.\n")

        return {
            'success': True,
            'message': f'Successfully processed {len(invoices)} invoice(s) across {current_week} week(s).',
            'invoice_count': len(invoices),
            'week_count': current_week
        }
    except Exception as e:
        print(f"Error during processing: {str(e)}")
        return {
            'success': False,
            'message': f'Processing error: {str(e)}'
        }
